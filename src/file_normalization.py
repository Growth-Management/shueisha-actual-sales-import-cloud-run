from __future__ import annotations

import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
import re
from typing import Any


CSV_CONTENT_TYPE = "text/csv; charset=utf-8"

PROVIDER_RULES = {
    "apple": {
        "columns": [
            "sales_yyyymm",
            "start_date",
            "end_date",
            "upc",
            "isrc_isbn",
            "vendor_identifier",
            "quantity",
            "partner_share",
            "extended_partner_share",
            "partner_share_currency",
            "sales_or_return",
            "apple_identifier",
            "artist_show_developer_author",
            "title",
            "label_studio_network_developer_publisher",
            "grid",
            "product_type_identifier",
            "isan_other_identifier",
            "country_of_sale",
            "pre_order_flag",
            "promo_code",
            "customer_price",
            "customer_currency",
            "currency",
            "exchange_rate",
            "deposit_amount_jpy",
            "sales_amount_jpy",
            "domestic_overseas",
            "app",
            "sales_category",
            "id_excerpt",
            "jdcn",
            "cp_price",
            "cp",
        ],
        "required_columns": [
            "start_date",
            "end_date",
            "vendor_identifier",
            "quantity",
            "extended_partner_share",
            "sales_or_return",
            "apple_identifier",
            "title",
            "product_type_identifier",
            "country_of_sale",
            "customer_price",
            "customer_currency",
            "deposit_amount_jpy",
            "sales_amount_jpy",
        ],
        "aliases": {
            "start_date": ["start date"],
            "end_date": ["end date"],
            "upc": ["upc"],
            "isrc_isbn": ["isrc/isbn", "isrc isbn"],
            "vendor_identifier": ["vendor identifier"],
            "quantity": ["quantity"],
            "partner_share": ["partner share"],
            "extended_partner_share": ["extended partner share"],
            "partner_share_currency": ["partner share currency"],
            "sales_or_return": ["sales or return"],
            "apple_identifier": ["apple identifier"],
            "artist_show_developer_author": ["artist/show/developer/author", "artist show developer author"],
            "title": ["title"],
            "label_studio_network_developer_publisher": [
                "label/studio/network/developer/publisher",
                "label studio network developer publisher",
            ],
            "grid": ["grid"],
            "product_type_identifier": ["product type identifier"],
            "isan_other_identifier": ["isan/other identifier", "isan other identifier"],
            "country_of_sale": ["country of sale"],
            "pre_order_flag": ["pre-order flag", "pre order flag"],
            "promo_code": ["promo code"],
            "customer_price": ["customer price"],
            "customer_currency": ["customer currency"],
            "currency": ["通貨", "currency"],
            "exchange_rate": ["レート", "exchange rate"],
            "deposit_amount_jpy": ["入金額（円）", "入金額(円)", "deposit amount jpy"],
            "sales_amount_jpy": ["売上額（円）", "売上額(円)", "sales amount jpy"],
            "domestic_overseas": ["国内／海外", "国内/海外"],
            "app": ["アプリ", "app"],
            "sales_category": ["種別", "sales category"],
            "id_excerpt": ["id抜粋", "id excerpt"],
            "jdcn": ["jdcn"],
            "cp_price": ["cp価格", "cp price"],
            "cp": ["cp"],
        },
        "types": {
            "start_date": "date",
            "end_date": "date",
            "quantity": "integer",
            "partner_share": "decimal",
            "extended_partner_share": "decimal",
            "customer_price": "decimal",
            "exchange_rate": "decimal",
            "deposit_amount_jpy": "decimal",
            "sales_amount_jpy": "decimal",
            "cp_price": "decimal",
            "cp": "decimal",
        },
    },
    "googleplay": {
        "columns": [
            "sales_yyyymm",
            "description",
            "transaction_date",
            "transaction_time",
            "tax_type",
            "transaction_type",
            "refund_type",
            "product_title",
            "package_id",
            "product_type",
            "sku_id",
            "hardware",
            "buyer_country",
            "buyer_state",
            "buyer_postal_code",
            "buyer_currency",
            "amount_buyer_currency",
            "currency_conversion_rate",
            "merchant_currency",
            "amount_merchant_currency",
            "fee_category",
            "sales_category",
            "content_category",
            "jdcn",
            "download_count",
            "sales_unit_price",
            "fee",
            "deposit_amount",
            "cp_price",
            "cp",
            "base_price",
            "tax",
        ],
        "required_columns": [
            "description",
            "transaction_date",
            "product_title",
            "package_id",
            "sku_id",
            "buyer_currency",
            "amount_buyer_currency",
            "merchant_currency",
            "amount_merchant_currency",
            "sales_category",
            "download_count",
            "sales_unit_price",
            "deposit_amount",
            "base_price",
        ],
        "aliases": {
            "transaction_date": ["transaction_date", "transaction date", "date", "日付"],
            "description": ["description"],
            "transaction_time": ["transaction time"],
            "tax_type": ["tax type"],
            "transaction_type": ["transaction type"],
            "refund_type": ["refund type"],
            "product_title": ["product title"],
            "package_id": ["package id", "package_name", "package name", "package", "app package", "アプリパッケージ"],
            "product_type": ["product type"],
            "sku_id": ["sku id", "sku", "product id", "product_id", "商品id", "商品ID"],
            "hardware": ["hardware"],
            "buyer_country": ["buyer country"],
            "buyer_state": ["buyer state"],
            "buyer_postal_code": ["buyer postal code"],
            "buyer_currency": ["buyer currency"],
            "amount_buyer_currency": ["amount (buyer currency)", "amount buyer currency"],
            "currency_conversion_rate": ["currency conversion rate"],
            "merchant_currency": ["merchant currency"],
            "amount_merchant_currency": [
                "amount (merchant currency)",
                "amount merchant currency",
                "developer_proceeds",
                "developer proceeds",
            ],
            "fee_category": ["手数料区分", "fee category"],
            "sales_category": ["売上区分", "sales category"],
            "content_category": ["コンテンツ区分", "content category"],
            "jdcn": ["jdcn"],
            "download_count": ["dl数", "download count"],
            "sales_unit_price": ["売上単価", "sales unit price"],
            "fee": ["手数料", "fee"],
            "deposit_amount": ["入金", "deposit amount"],
            "cp_price": ["cp価格", "cp price"],
            "cp": ["cp"],
            "base_price": ["本体価", "base price"],
            "tax": ["税", "tax"],
        },
        "types": {
            "transaction_date": "date",
            "amount_buyer_currency": "decimal",
            "currency_conversion_rate": "decimal",
            "amount_merchant_currency": "decimal",
            "download_count": "integer",
            "sales_unit_price": "decimal",
            "fee": "decimal",
            "deposit_amount": "decimal",
            "cp_price": "decimal",
            "cp": "decimal",
            "base_price": "decimal",
            "tax": "decimal",
        },
    },
}


def normalize_drive_file(
    *,
    provider: str,
    sales_yyyymm: str | None,
    file_name: str,
    mime_type: str | None,
    data: bytes,
) -> dict[str, Any]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv" or (mime_type or "").lower() == "text/csv":
        normalized_data = _normalize_rows(
            provider=provider,
            sales_yyyymm=sales_yyyymm,
            rows=_csv_rows(data),
        )
        return {
            "file_name": file_name,
            "content_type": CSV_CONTENT_TYPE,
            **normalized_data,
            "format": "csv",
            "was_converted": False,
        }
    if suffix in {".xlsx", ".xlsm"}:
        normalized_data = _normalize_rows(
            provider=provider,
            sales_yyyymm=sales_yyyymm,
            rows=_xlsx_rows(data),
        )
        return {
            "file_name": f"{Path(file_name).stem}.csv",
            "content_type": CSV_CONTENT_TYPE,
            **normalized_data,
            "format": "csv",
            "was_converted": True,
        }
    raise ValueError(f"unsupported file format for BigQuery load: {file_name}")


def _normalize_rows(*, provider: str, sales_yyyymm: str | None, rows: list[list[Any]]) -> dict[str, Any]:
    rule = PROVIDER_RULES.get(provider)
    if rule is None:
        raise ValueError("provider must be apple or googleplay")

    header_index, column_map = _find_header(rows, rule)
    body_rows = [
        _normalized_record(
            row=row,
            sales_yyyymm=sales_yyyymm,
            rule=rule,
            column_map=column_map,
        )
        for row in rows[header_index + 1 :]
        if _has_values(row)
    ]

    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=rule["columns"], lineterminator="\n")
    writer.writeheader()
    for record in body_rows:
        writer.writerow(record)
    return {
        "data": output.getvalue().encode("utf-8"),
        "columns": rule["columns"],
        "row_count": len(body_rows),
        "header_row_index": header_index,
    }


def _find_header(rows: list[list[Any]], rule: dict[str, Any]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows):
        normalized_cells = [_normalize_header(value) for value in row]
        column_map: dict[str, int] = {}
        for canonical, aliases in rule["aliases"].items():
            alias_set = {_normalize_header(alias) for alias in aliases}
            for column_index, cell in enumerate(normalized_cells):
                if cell in alias_set:
                    column_map[canonical] = column_index
                    break
        missing = [column for column in rule["required_columns"] if column not in column_map]
        if not missing:
            return index, column_map
    raise ValueError(f"required columns are missing: {', '.join(rule['required_columns'])}")


def _normalized_record(
    *,
    row: list[Any],
    sales_yyyymm: str | None,
    rule: dict[str, Any],
    column_map: dict[str, int],
) -> dict[str, Any]:
    record: dict[str, Any] = {"sales_yyyymm": sales_yyyymm or ""}
    for column in rule["columns"]:
        if column == "sales_yyyymm":
            continue
        value = row[column_map[column]] if column in column_map and column_map[column] < len(row) else None
        record[column] = _coerce_value(value, rule["types"].get(column))
    return record


def _coerce_value(value: Any, value_type: str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
    if value == "":
        return ""
    if value_type == "date":
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%Y%m%d", "%b %d %Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass
        return text
    if value_type == "integer":
        text = re.sub(r"[, ]", "", str(value).strip())
        if text.startswith("(") and text.endswith(")"):
            text = f"-{text[1:-1]}"
        try:
            return str(int(Decimal(text)))
        except InvalidOperation:
            return str(value).strip()
    if value_type == "decimal":
        text = re.sub(r"[,¥$ ]", "", str(value).strip())
        if text.startswith("(") and text.endswith(")"):
            text = f"-{text[1:-1]}"
        try:
            return str(Decimal(text))
        except InvalidOperation:
            return str(value).strip()
    return str(value).strip()


def _csv_rows(data: bytes) -> list[list[Any]]:
    text = _decode_csv(data)
    return [row for row in csv.reader(StringIO(text))]


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def _xlsx_rows(data: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    try:
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s_\-]+", " ", str(value or "").strip().lower())


def _has_values(row: list[Any]) -> bool:
    return any(str(value).strip() for value in row if value is not None)
